#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
分析E:\gujia\assets\sources\baobao.xlsx的结构
"""

import pandas as pd
import openpyxl
import sys
import os

def analyze_excel(file_path):
    print(f"分析文件: {file_path}")
    print("文件大小:", os.path.getsize(file_path), "bytes")

    # 使用openpyxl读取工作表信息
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    print("\n=== 工作表列表 ===")
    for i, sheet_name in enumerate(wb.sheetnames):
        print(f"{i+1}. {sheet_name}")

    # 检查第一个工作表
    sheet_name = wb.sheetnames[0]
    print(f"\n=== 详细分析工作表: '{sheet_name}' ===")

    ws = wb[sheet_name]

    # 读取前30行，前20列的数据
    data = []
    max_row = min(ws.max_row, 50)
    max_col = min(ws.max_column, 20)

    print(f"工作表尺寸: {ws.max_row} 行, {ws.max_column} 列")
    print(f"读取前 {max_row} 行, 前 {max_col} 列")

    for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col, values_only=True):
        data.append(row)

    # 显示表头行（前10行）
    print("\n=== 前10行数据 ===")
    for i, row in enumerate(data[:10]):
        print(f"行 {i+1}: {row}")

    # 尝试用pandas读取（可能需要指定表头行）
    print("\n=== 尝试用pandas读取 ===")
    try:
        # 尝试多种可能的表头行
        for header_row in [0, 1, 2, 3]:
            print(f"\n尝试表头行={header_row}:")
            df = pd.read_excel(file_path, sheet_name=0, header=header_row, nrows=20)
            print(f"形状: {df.shape}")
            print("列名:", list(df.columns))
            print("前3行:")
            print(df.head(3))
            print("-" * 50)
    except Exception as e:
        print(f"pandas读取失败: {e}")

    wb.close()
    print("\n=== 分析完成 ===")

if __name__ == "__main__":
    excel_file = "../E:\gujia\assets\sources\baobao.xlsx"
    if not os.path.exists(excel_file):
        print(f"文件不存在: {excel_file}")
        sys.exit(1)

    analyze_excel(excel_file)