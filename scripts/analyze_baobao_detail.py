#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
详细分析E:\gujia\assets\sources\baobao.xlsx的结构
"""

import openpyxl
import sys
import os

def analyze_excel_detail(file_path):
    print(f"分析文件: {file_path}")

    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    sheet_name = wb.sheetnames[0]
    print(f"工作表: {sheet_name}")
    ws = wb[sheet_name]

    print(f"总行数: {ws.max_row}, 总列数: {ws.max_column}")

    # 收集所有非空行
    non_empty_rows = []
    for r in range(1, ws.max_row + 1):
        row_vals = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
        # 检查是否有非空值
        if any(cell is not None for cell in row_vals):
            non_empty_rows.append((r, row_vals))

    print(f"\n非空行数: {len(non_empty_rows)}")

    # 打印前50行非空行
    print("\n=== 前50行非空数据 ===")
    for i, (r, row_vals) in enumerate(non_empty_rows[:50]):
        # 清理显示：None显示为空字符串
        row_display = []
        for v in row_vals:
            if v is None:
                row_display.append('')
            elif isinstance(v, (int, float)):
                row_display.append(str(v))
            else:
                row_display.append(str(v).strip())

        print(f"行 {r:3d}: {row_display}")

    # 分析列结构：检查每列的数据类型
    print("\n=== 列分析 ===")
    for c in range(1, min(ws.max_column, 15) + 1):
        col_vals = []
        for r in range(1, min(ws.max_row, 100) + 1):
            val = ws.cell(r, c).value
            if val is not None:
                col_vals.append(val)

        if col_vals:
            # 统计类型
            types = {}
            for v in col_vals:
                t = type(v).__name__
                types[t] = types.get(t, 0) + 1

            sample = col_vals[0] if len(col_vals) > 0 else None
            print(f"列 {c:2d}: 非空值 {len(col_vals)} 个, 类型分布: {types}, 示例: {sample}")

    # 尝试检测分组模式
    print("\n=== 分组模式分析 ===")
    current_battle_level = None
    current_pet_type = None
    current_pet_quality = None

    for r in range(1, min(ws.max_row, 200) + 1):
        battle_level = ws.cell(r, 1).value
        pet_type = ws.cell(r, 2).value
        pet_quality = ws.cell(r, 3).value

        if battle_level is not None:
            current_battle_level = battle_level
        if pet_type is not None:
            current_pet_type = pet_type
        if pet_quality is not None:
            current_pet_quality = pet_quality

        skill = ws.cell(r, 4).value
        if skill is not None and skill != '':
            # 收集技能价格
            prices = []
            for c in range(5, ws.max_column + 1):
                price = ws.cell(r, c).value
                if price is not None:
                    prices.append((c-4, price))  # c-4得到技能数量

            if prices:
                print(f"行 {r:3d}: 战斗等级={current_battle_level}, 类型={current_pet_type}, 资质={current_pet_quality}, 技能={skill}, 价格: {prices}")

    wb.close()

if __name__ == "__main__":
    excel_file = r"E:\gujia\assets\sources\baobao.xlsx"
    if not os.path.exists(excel_file):
        print(f"文件不存在: {excel_file}")
        sys.exit(1)

    analyze_excel_detail(excel_file)