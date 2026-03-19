#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
分析宝宝价格表模式
"""

import openpyxl
import sys
import os

def analyze_pattern(file_path):
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    sheet_name = wb.sheetnames[0]
    ws = wb[sheet_name]

    current_battle_level = None
    current_pet_type = None
    current_pet_quality = None

    print("行号\t战斗等级\t宝宝类型\t宝宝资质\t技能\t价格(4-13技能)")

    for r in range(1, min(ws.max_row, 200) + 1):
        col1 = ws.cell(r, 1).value
        col2 = ws.cell(r, 2).value
        col3 = ws.cell(r, 3).value
        col4 = ws.cell(r, 4).value

        # 更新状态
        if col1 is not None and str(col1).strip() != '' and str(col1).strip() != '战斗等级':
            current_battle_level = str(col1).strip()
            # 新战斗等级，重置类型和资质（可能变化）
            current_pet_type = None
            current_pet_quality = None

        if col2 is not None and str(col2).strip() != '' and str(col2).strip() != '宝宝类型':
            current_pet_type = str(col2).strip()

        if col3 is not None and str(col3).strip() != '' and str(col3).strip() != '宝宝资质':
            current_pet_quality = str(col3).strip()

        # 检查是否是技能行
        if col4 is None or str(col4).strip() == '' or str(col4).strip() == '技能':
            continue

        skill = str(col4).strip()

        # 收集价格
        prices = []
        for c in range(5, 15):  # 4-13技能
            price = ws.cell(r, c).value
            if price is not None:
                prices.append(f"{c-4}:{price}")
            else:
                prices.append(f"{c-4}:")

        price_str = " ".join(prices)

        print(f"{r}\t{current_battle_level}\t{current_pet_type}\t{current_pet_quality}\t{skill}\t{price_str}")

    wb.close()

if __name__ == "__main__":
    excel_file = r"E:\gujia\assets\sources\baobao.xlsx"
    analyze_pattern(excel_file)