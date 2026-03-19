#!/usr/bin/env python3
# -*- coding: utf-8 -*-
import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8')

def analyze_patterns(path):
    print(f"分析灵饰数据模式: {path}")

    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb['工作表2']  # 直接使用工作表2

    # 检查第5列和第9列（部位列）是否有值
    print("检查'部位'列（第5列和第9列）:")
    row_count = 0
    for r in range(1, min(51, ws.max_row + 1)):  # 检查前50行
        col5 = ws.cell(r, 5).value  # 第5列：部位
        col9 = ws.cell(r, 9).value  # 第9列：部位

        if col5 or col9:
            print(f"  行{r}: 第5列={repr(col5)}, 第9列={repr(col9)}")
            row_count += 1

    if row_count == 0:
        print("  前50行中'部位'列都为空")

    # 分析数据块模式
    print("\n分析数据块模式:")
    current_position = None
    current_category = None
    current_level = None
    block_start = 0

    for r in range(2, min(100, ws.max_row + 1)):  # 从第2行开始
        position = ws.cell(r, 1).value  # 第1列：灵饰部位
        category = ws.cell(r, 2).value  # 第2列：灵饰类别
        level = ws.cell(r, 3).value    # 第3列：灵饰等级

        if position or category or (level is not None):
            if position != current_position or category != current_category or level != current_level:
                if current_position is not None:
                    print(f"  块 {block_start}-{r-1}: {current_position}/{current_category}/{current_level}")
                current_position = position
                current_category = category
                current_level = level
                block_start = r

    if current_position is not None:
        print(f"  块 {block_start}-{min(100, ws.max_row)}: {current_position}/{current_category}/{current_level}")

    # 检查左属性和右属性的模式
    print("\n检查属性模式（前20行）:")
    for r in range(2, min(22, ws.max_row + 1)):
        left_attr = ws.cell(r, 6).value  # 第6列：属性组合（左）
        right_attr = ws.cell(r, 10).value  # 第10列：属性组合（右）

        if left_attr or right_attr:
            print(f"  行{r}: 左={repr(left_attr)}, 右={repr(right_attr)}")

    # 尝试推断类型
    print("\n尝试推断类型规则:")
    print("假设1: '前排' -> 戒指和耳饰")
    print("假设2: '后排' -> 配饰和手镯")
    print("可能需要根据属性内容或行列位置推断具体类型")

if __name__ == "__main__":
    excel_path = r"E:\gujia\assets\sources\lingshi.xlsx"
    analyze_patterns(excel_path)