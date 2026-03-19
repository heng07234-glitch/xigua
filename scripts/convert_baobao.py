#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
转换E:\gujia\assets\sources\baobao.xlsx为JS数据
"""

import openpyxl
import json
import re
import sys
import os

sys.stdout.reconfigure(encoding='utf-8')

def parse_price(price_val):
    """解析价格，返回数字或None"""
    if price_val is None:
        return None
    if isinstance(price_val, (int, float)):
        # 如果数字很小可能是万单位？但宝宝价格看起来是实际数值
        return float(price_val)
    if isinstance(price_val, str):
        price_str = price_val.strip()
        if price_str == '':
            return None
        # 尝试提取数字
        match = re.search(r'[\d.]+', price_str)
        if match:
            try:
                return float(match.group())
            except:
                return None
    return None

def convert_excel_to_js(excel_path, output_js_path):
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    sheet_name = wb.sheetnames[0]
    ws = wb[sheet_name]

    all_data = []

    # 当前分组状态
    current_battle_level = None
    current_pet_type = None
    current_pet_quality = None

    # 遍历所有行
    for r in range(1, ws.max_row + 1):
        # 读取前4列的关键字段
        battle_level = ws.cell(r, 1).value
        pet_type = ws.cell(r, 2).value
        pet_quality = ws.cell(r, 3).value
        skill = ws.cell(r, 4).value

        # 更新当前状态
        if battle_level is not None and str(battle_level).strip() != '':
            current_battle_level = str(battle_level).strip()
        if pet_type is not None and str(pet_type).strip() != '':
            current_pet_type = str(pet_type).strip()
        if pet_quality is not None and str(pet_quality).strip() != '':
            current_pet_quality = str(pet_quality).strip()

        # 如果技能列为空，跳过
        if skill is None or str(skill).strip() == '':
            continue

        skill_str = str(skill).strip()

        # 检查是否为表头行（技能列包含"技能"字样）
        if '技能' in skill_str:
            continue

        # 检查是否为难度系数行（包含数字和小数点）
        if skill_str.replace('.', '', 1).isdigit():
            # 这是难度系数行，可能用于整体乘数，暂时跳过
            continue

        # 从第5列开始读取技能数量价格（4技能到13技能）
        for skill_col in range(5, 15):  # 对应Excel列E到N（4-13技能）
            price = ws.cell(r, skill_col).value
            parsed_price = parse_price(price)

            if parsed_price is None:
                continue

            # 计算技能数量（列索引减去4得到技能数）
            skill_count = skill_col - 4  # 5->1, 但应该是4技能开始？实际列对应：5列是4技能，6列是5技能...
            # 第5列是"4技能"，所以技能数量 = (列索引 - 5) + 4
            skill_count = (skill_col - 5) + 4

            # 构建数据记录
            record = {
                "battleLevel": current_battle_level,
                "petType": current_pet_type,
                "petQuality": current_pet_quality,
                "skill": skill_str,
                "skillCount": skill_count,
                "price": parsed_price
            }
            all_data.append(record)

    print(f"解析出 {len(all_data)} 条宝宝价格记录")

    # 统计数据
    battle_levels = set([d["battleLevel"] for d in all_data])
    pet_types = set([d["petType"] for d in all_data])
    pet_qualities = set([d["petQuality"] for d in all_data])
    skills = set([d["skill"] for d in all_data])

    print(f"战斗等级: {sorted(battle_levels)}")
    print(f"宝宝类型: {sorted(pet_types)}")
    print(f"宝宝资质: {sorted(pet_qualities)}")
    print(f"技能数量: {len(skills)} 种")
    print(f"技能示例: {list(skills)[:10]}")

    # 生成JS文件
    js_content = f"""// 宝宝价格规则库
// 自动生成自: {os.path.basename(excel_path)}
// 生成时间: {sys.version}
// 记录数: {len(all_data)}

const BAOBAO_DATA = {json.dumps(all_data, ensure_ascii=False, indent=2)};

// 导出供其他模块使用
if (typeof module !== 'undefined' && module.exports) {{
    module.exports = BAOBAO_DATA;
}}
"""

    with open(output_js_path, 'w', encoding='utf-8') as f:
        f.write(js_content)

    print(f"\n已生成JS数据文件: {output_js_path}")

    # 打印示例
    print("\n前5条记录示例:")
    for i, record in enumerate(all_data[:5]):
        print(f"  {i+1}: {record}")

    wb.close()

if __name__ == "__main__":
    excel_path = r"E:\gujia\assets\sources\baobao.xlsx"
    if not os.path.exists(excel_path):
        print(f"文件不存在: {excel_path}")
        sys.exit(1)

    output_path = r"E:\gujia\assets\data\baobao_data.js"
    convert_excel_to_js(excel_path, output_path)