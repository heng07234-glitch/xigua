#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
转换E:\gujia\assets\sources\baobao.xlsx为完整JS数据（包括空价格）
"""

import openpyxl
import json
import re
import sys
import os

sys.stdout.reconfigure(encoding='utf-8')

def parse_price(price_val):
    """解析价格，返回数字或None"""
    if price_val is None:
        return None
    if isinstance(price_val, (int, float)):
        return float(price_val)
    if isinstance(price_val, str):
        price_str = price_val.strip()
        if price_str == '':
            return None
        # 检查是否为技能数量表头（如"4技能"）
        if '技能' in price_str:
            return None
        # 尝试提取数字
        match = re.search(r'[\d.]+', price_str)
        if match:
            try:
                return float(match.group())
            except:
                return None
    return None

def convert_excel_to_js(excel_path, output_js_path):
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    sheet_name = wb.sheetnames[0]
    ws = wb[sheet_name]

    all_data = []
    multipliers = []

    current_battle_level = None
    current_pet_type = None
    current_pet_quality = None

    # 遍历所有行
    for r in range(1, ws.max_row + 1):
        col1 = ws.cell(r, 1).value
        col2 = ws.cell(r, 2).value
        col3 = ws.cell(r, 3).value
        col4 = ws.cell(r, 4).value

        # 处理空值
        col1_str = str(col1).strip() if col1 is not None else ''
        col2_str = str(col2).strip() if col2 is not None else ''
        col3_str = str(col3).strip() if col3 is not None else ''
        col4_str = str(col4).strip() if col4 is not None else ''

        # 跳过完全空的行
        if not col1_str and not col2_str and not col3_str and not col4_str:
            continue

        # 更新状态
        if col1_str and col1_str != '战斗等级' and not col1_str.startswith('4技能'):
            current_battle_level = col1_str
            # 新战斗等级，不重置类型和资质，因为可能延续

        if col2_str and col2_str != '宝宝类型':
            current_pet_type = col2_str

        if col3_str and col3_str != '宝宝资质':
            current_pet_quality = col3_str

        # 检查是否为技能行（第4列有值且不是表头）
        if not col4_str or col4_str == '技能':
            continue

        # 检查是否为难度系数（纯数字）
        if col4_str.replace('.', '', 1).isdigit():
            try:
                multiplier = float(col4_str)
                multipliers.append({
                    "battleLevel": current_battle_level,
                    "petType": current_pet_type,
                    "petQuality": current_pet_quality,
                    "multiplier": multiplier
                })
            except:
                pass
            continue

        skill_name = col4_str

        # 为每个技能数量创建记录
        for skill_col in range(5, 15):
            price = ws.cell(r, skill_col).value
            parsed_price = parse_price(price)

            skill_count = (skill_col - 5) + 4

            record = {
                "battleLevel": current_battle_level,
                "petType": current_pet_type,
                "petQuality": current_pet_quality if current_pet_quality else '',
                "skill": skill_name,
                "skillCount": skill_count,
                "price": parsed_price
            }
            all_data.append(record)

    print(f"解析出 {len(all_data)} 条宝宝价格记录（含空价格）")
    print(f"解析出 {len(multipliers)} 条难度系数记录")

    # 过滤掉完全空价格的记录（可选）
    non_empty_data = [d for d in all_data if d["price"] is not None]
    print(f"其中非空价格记录: {len(non_empty_data)} 条")

    # 统计数据
    battle_levels = set([d["battleLevel"] for d in all_data if d["battleLevel"]])
    pet_types = set([d["petType"] for d in all_data if d["petType"]])
    pet_qualities = set([d["petQuality"] for d in all_data if d["petQuality"]])
    skills = set([d["skill"] for d in all_data])

    print(f"战斗等级: {sorted(battle_levels)}")
    print(f"宝宝类型: {sorted(pet_types)}")
    print(f"宝宝资质: {sorted(pet_qualities)}")
    print(f"技能种类: {len(skills)}")
    print(f"技能示例: {list(skills)[:20]}")

    # 生成JS文件
    js_content = f"""// 宝宝价格规则库（完整数据，含空价格）
// 自动生成自: {os.path.basename(excel_path)}
// 生成时间: {sys.version}
// 总记录数: {len(all_data)}
// 非空价格记录: {len(non_empty_data)}
// 难度系数: {len(multipliers)}

const BAOBAO_DATA = {json.dumps(all_data, ensure_ascii=False, indent=2)};

// 难度系数（用于价格调整）
const BAOBAO_MULTIPLIERS = {json.dumps(multipliers, ensure_ascii=False, indent=2)};

// 导出供其他模块使用
if (typeof module !== 'undefined' && module.exports) {{
    module.exports = {{ BAOBAO_DATA, BAOBAO_MULTIPLIERS }};
}}
"""

    with open(output_js_path, 'w', encoding='utf-8') as f:
        f.write(js_content)

    print(f"\n已生成JS数据文件: {output_js_path}")

    wb.close()

if __name__ == "__main__":
    excel_path = r"E:\gujia\assets\sources\baobao.xlsx"
    if not os.path.exists(excel_path):
        print(f"文件不存在: {excel_path}")
        sys.exit(1)

    output_path = r"E:\gujia\assets\data\baobao_data_full.js"
    convert_excel_to_js(excel_path, output_path)