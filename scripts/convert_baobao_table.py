#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
转换宝宝价格表为表格格式JS数据
每个技能一行，包含4-13技能价格
"""

import openpyxl
import json
import re
import sys
import os

sys.stdout.reconfigure(encoding='utf-8')

def parse_price(price_val):
    if price_val is None:
        return None
    if isinstance(price_val, (int, float)):
        return float(price_val)
    if isinstance(price_val, str):
        price_str = price_val.strip()
        if price_str == '':
            return None
        if '技能' in price_str:
            return None
        match = re.search(r'[\d.]+', price_str)
        if match:
            try:
                return float(match.group())
            except:
                return None
    return None

def convert_excel_to_js(excel_path, output_js_path):
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    sheet_name = wb.sheetnames[0]
    ws = wb[sheet_name]

    table_data = []  # 每个技能一行
    multipliers = []

    current_battle_level = None
    current_pet_type = None
    current_pet_quality = None

    # 首先收集所有技能行
    skill_rows = []

    for r in range(1, ws.max_row + 1):
        col1 = ws.cell(r, 1).value
        col2 = ws.cell(r, 2).value
        col3 = ws.cell(r, 3).value
        col4 = ws.cell(r, 4).value

        col1_str = str(col1).strip() if col1 is not None else ''
        col2_str = str(col2).strip() if col2 is not None else ''
        col3_str = str(col3).strip() if col3 is not None else ''
        col4_str = str(col4).strip() if col4 is not None else ''

        if not col1_str and not col2_str and not col3_str and not col4_str:
            continue

        # 更新状态
        if col1_str and col1_str != '战斗等级' and not col1_str.startswith('4技能'):
            current_battle_level = col1_str

        if col2_str and col2_str != '宝宝类型':
            current_pet_type = col2_str

        if col3_str and col3_str != '宝宝资质':
            current_pet_quality = col3_str

        if not col4_str or col4_str == '技能':
            continue

        # 难度系数
        if col4_str.replace('.', '', 1).isdigit():
            try:
                multiplier = float(col4_str)
                multipliers.append({
                    "battleLevel": current_battle_level,
                    "petType": current_pet_type,
                    "petQuality": current_pet_quality,
                    "multiplier": multiplier
                })
            except:
                pass
            continue

        skill_name = col4_str

        # 收集价格
        prices = {}
        for skill_col in range(5, 15):
            price = ws.cell(r, skill_col).value
            parsed_price = parse_price(price)
            skill_count = (skill_col - 5) + 4
            prices[str(skill_count)] = parsed_price

        # 添加到技能行列表
        skill_rows.append({
            "row": r,
            "battleLevel": current_battle_level,
            "petType": current_pet_type,
            "petQuality": current_pet_quality if current_pet_quality else '',
            "skill": skill_name,
            "prices": prices
        })

    # 合并相同技能的行（可能有多行？）
    # 按战斗等级、类型、资质、技能分组
    grouped = {}
    for row in skill_rows:
        key = (row["battleLevel"], row["petType"], row["petQuality"], row["skill"])
        if key not in grouped:
            grouped[key] = row
        else:
            # 合并价格（后行的价格覆盖前行）
            existing = grouped[key]
            for skill_count, price in row["prices"].items():
                if price is not None:
                    existing["prices"][skill_count] = price

    table_data = list(grouped.values())

    print(f"解析出 {len(table_data)} 个技能行")
    print(f"解析出 {len(multipliers)} 条难度系数记录")

    # 统计数据
    battle_levels = set([d["battleLevel"] for d in table_data])
    pet_types = set([d["petType"] for d in table_data])
    pet_qualities = set([d["petQuality"] for d in table_data if d["petQuality"]])
    skills = set([d["skill"] for d in table_data])

    print(f"战斗等级: {sorted(battle_levels)}")
    print(f"宝宝类型: {sorted(pet_types)}")
    print(f"宝宝资质: {sorted(pet_qualities)}")
    print(f"技能种类: {len(skills)}")
    print(f"技能列表: {sorted(skills)}")

    # 生成JS文件
    js_content = f"""// 宝宝价格表（表格格式）
// 自动生成自: {os.path.basename(excel_path)}
// 生成时间: {sys.version}
// 技能行数: {len(table_data)}
// 难度系数: {len(multipliers)}

const BAOBAO_TABLE_DATA = {json.dumps(table_data, ensure_ascii=False, indent=2)};

const BAOBAO_MULTIPLIERS = {json.dumps(multipliers, ensure_ascii=False, indent=2)};

// 导出
if (typeof module !== 'undefined' && module.exports) {{
    module.exports = {{ BAOBAO_TABLE_DATA, BAOBAO_MULTIPLIERS }};
}}
"""

    with open(output_js_path, 'w', encoding='utf-8') as f:
        f.write(js_content)

    print(f"\n已生成JS数据文件: {output_js_path}")

    # 打印示例
    print("\n前3条记录示例:")
    for i, row in enumerate(table_data[:3]):
        print(f"  {i+1}: 战斗等级={row['battleLevel']}, 类型={row['petType']}, 资质={row['petQuality']}, 技能={row['skill']}")
        print(f"     价格: {row['prices']}")

    wb.close()

if __name__ == "__main__":
    excel_path = r"E:\gujia\assets\sources\baobao.xlsx"
    if not os.path.exists(excel_path):
        print(f"文件不存在: {excel_path}")
        sys.exit(1)

    output_path = r"E:\gujia\assets\data\baobao_table_data.js"
    convert_excel_to_js(excel_path, output_path)