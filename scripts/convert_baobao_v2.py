#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
转换E:\gujia\assets\sources\baobao.xlsx为JS数据 - 版本2
"""

import openpyxl
import json
import re
import sys
import os

sys.stdout.reconfigure(encoding='utf-8')

def parse_price(price_val):
    """解析价格，返回数字或None"""
    if price_val is None:
        return None
    if isinstance(price_val, (int, float)):
        return float(price_val)
    if isinstance(price_val, str):
        price_str = price_val.strip()
        if price_str == '':
            return None
        # 尝试提取数字
        match = re.search(r'[\d.]+', price_str)
        if match:
            try:
                return float(match.group())
            except:
                return None
    return None

def convert_excel_to_js(excel_path, output_js_path):
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    sheet_name = wb.sheetnames[0]
    ws = wb[sheet_name]

    all_data = []
    multipliers = []  # 存储难度系数

    # 当前分组状态
    current_battle_level = None
    current_pet_type = None
    current_pet_quality = None

    # 遍历所有行
    for r in range(1, ws.max_row + 1):
        # 读取前4列
        col1 = ws.cell(r, 1).value
        col2 = ws.cell(r, 2).value
        col3 = ws.cell(r, 3).value
        col4 = ws.cell(r, 4).value

        # 处理空值
        col1_str = str(col1).strip() if col1 is not None else ''
        col2_str = str(col2).strip() if col2 is not None else ''
        col3_str = str(col3).strip() if col3 is not None else ''
        col4_str = str(col4).strip() if col4 is not None else ''

        # 重置逻辑：如果第1列有非空值且不是表头，则是新的战斗等级
        if col1_str and col1_str != '战斗等级' and not col1_str.startswith('4技能'):
            current_battle_level = col1_str
            # 新战斗等级组，重置类型和资质
            current_pet_type = None
            current_pet_quality = None

        # 第2列：宝宝类型
        if col2_str and col2_str != '宝宝类型':
            current_pet_type = col2_str

        # 第3列：宝宝资质
        if col3_str and col3_str != '宝宝资质':
            current_pet_quality = col3_str

        # 第4列：技能或难度系数
        if not col4_str or col4_str == '技能':
            continue

        # 检查是否为难度系数（纯数字，可能带小数点）
        if col4_str.replace('.', '', 1).isdigit():
            # 存储难度系数
            try:
                multiplier = float(col4_str)
                # 查找对应的类型和资质（可能在前面的行）
                multiplier_data = {
                    "battleLevel": current_battle_level,
                    "petType": current_pet_type,
                    "petQuality": current_pet_quality,
                    "multiplier": multiplier
                }
                multipliers.append(multiplier_data)
                print(f"找到难度系数: {multiplier_data}")
            except:
                pass
            continue

        # 现在col4_str是技能名称
        skill_name = col4_str

        # 读取技能数量价格（第5-14列，对应4-13技能）
        for skill_col in range(5, 15):
            price = ws.cell(r, skill_col).value
            parsed_price = parse_price(price)

            if parsed_price is None:
                continue

            # 计算技能数量
            skill_count = (skill_col - 5) + 4  # 5->4, 6->5, ..., 14->13

            # 构建记录
            record = {
                "battleLevel": current_battle_level,
                "petType": current_pet_type,
                "petQuality": current_pet_quality if current_pet_quality else '',
                "skill": skill_name,
                "skillCount": skill_count,
                "price": parsed_price
            }
            all_data.append(record)

    print(f"解析出 {len(all_data)} 条宝宝价格记录")
    print(f"解析出 {len(multipliers)} 条难度系数记录")

    # 统计数据
    battle_levels = set([d["battleLevel"] for d in all_data])
    pet_types = set([d["petType"] for d in all_data if d["petType"]])
    pet_qualities = set([d["petQuality"] for d in all_data if d["petQuality"]])
    skills = set([d["skill"] for d in all_data])

    print(f"战斗等级: {sorted(battle_levels)}")
    print(f"宝宝类型: {sorted(pet_types)}")
    print(f"宝宝资质: {sorted(pet_qualities)}")
    print(f"技能数量: {len(skills)} 种")
    print(f"技能列表: {sorted(skills)}")

    # 生成JS文件
    js_content = f"""// 宝宝价格规则库
// 自动生成自: {os.path.basename(excel_path)}
// 生成时间: {sys.version}
// 记录数: {len(all_data)}
// 难度系数: {len(multipliers)}

const BAOBAO_DATA = {json.dumps(all_data, ensure_ascii=False, indent=2)};

// 难度系数（用于价格调整）
const BAOBAO_MULTIPLIERS = {json.dumps(multipliers, ensure_ascii=False, indent=2)};

// 导出供其他模块使用
if (typeof module !== 'undefined' && module.exports) {{
    module.exports = {{ BAOBAO_DATA, BAOBAO_MULTIPLIERS }};
}}
"""

    with open(output_js_path, 'w', encoding='utf-8') as f:
        f.write(js_content)

    print(f"\n已生成JS数据文件: {output_js_path}")

    # 打印示例
    print("\n前5条记录示例:")
    for i, record in enumerate(all_data[:5]):
        print(f"  {i+1}: {record}")

    wb.close()

if __name__ == "__main__":
    excel_path = r"E:\gujia\assets\sources\baobao.xlsx"
    if not os.path.exists(excel_path):
        print(f"文件不存在: {excel_path}")
        sys.exit(1)

    output_path = r"E:\gujia\assets\data\baobao_data_v2.js"
    convert_excel_to_js(excel_path, output_path)