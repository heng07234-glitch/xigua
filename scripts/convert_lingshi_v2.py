#!/usr/bin/env python3
# -*- coding: utf-8 -*-
import openpyxl
import json
import re
import sys
import os

sys.stdout.reconfigure(encoding='utf-8')

def parse_price(price_val):
    """解析价格，处理可能的字符串格式"""
    if price_val is None:
        return None
    if isinstance(price_val, (int, float)):
        return price_val
    if isinstance(price_val, str):
        # 移除空格和单位
        price_str = price_val.strip()
        if price_str == '':
            return None
        # 尝试提取数字
        match = re.search(r'[\d.]+', price_str)
        if match:
            try:
                return float(match.group())
            except:
                return None
    return None

def extract_type_from_desc(desc):
    """从属性描述中提取装备类型（戒指、耳饰、配饰、手镯）
    支持格式: 1) "（戒指）伤害／3伤害"  2) "戒指/伤害／3伤害" """
    if not desc:
        return ''
    desc_str = str(desc)

    # 格式1: 括号格式 "（戒指）伤害／3伤害"
    bracket_match = re.search(r'^（([^）]+)）', desc_str)
    if bracket_match:
        type_part = bracket_match.group(1).strip()
        if type_part in ['戒指', '耳饰', '配饰', '手镯']:
            return type_part

    # 格式2: 斜杠格式 "戒指/伤害／3伤害"
    if '/' in desc_str:
        type_part = desc_str.split('/')[0].strip()
        if type_part in ['戒指', '耳饰', '配饰', '手镯']:
            return type_part

    # 格式3: 直接查找关键字
    if '戒指' in desc_str:
        return '戒指'
    elif '耳饰' in desc_str:
        return '耳饰'
    elif '配饰' in desc_str:
        return '配饰'
    elif '手镯' in desc_str:
        return '手镯'
    else:
        return ''

def clean_desc(desc):
    """清理属性描述，移除类型部分
    支持格式: 1) "（戒指）伤害／3伤害" -> "伤害／3伤害"
             2) "戒指/伤害／3伤害" -> "伤害／3伤害" """
    if not desc:
        return ''
    desc_str = str(desc)

    # 格式1: 移除括号类型
    cleaned = re.sub(r'^（[^）]+）', '', desc_str)
    if cleaned != desc_str:
        return cleaned.strip()

    # 格式2: 移除斜杠前的类型部分
    if '/' in desc_str:
        parts = desc_str.split('/', 1)
        if len(parts) > 1:
            return parts[1].strip()

    # 其他格式: 返回原描述
    return desc_str.strip()

def infer_type(position, category, raw_attribute):
    """根据部位、类别和原始属性推断灵饰类型"""
    if not position or not category:
        return ''

    # 根据游戏常识推断
    if position == '前排':
        if category == '物理':
            return '戒指'
        elif category == '法系':
            return '耳饰'
        elif category == '辅助':
            # 辅助前排根据属性推断：治疗、封印等可能是耳饰
            if raw_attribute and ('治疗' in raw_attribute or '封印' in raw_attribute or '速度' in raw_attribute):
                return '耳饰'
            else:
                return '戒指'
        elif category == '通用':
            return '戒指'  # 默认
    elif position == '后排':
        if category == '辅助' or category == '通用':
            # 后排根据属性推断：气血、防御等
            if raw_attribute and ('气血' in raw_attribute or '防御' in raw_attribute):
                return '配饰'  # 配饰通常加气血防御
            else:
                return '手镯'
        else:
            return '配饰'  # 默认

    return ''

def convert_excel_to_js(path, output_js_path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]

    all_data = []
    current_position = None
    current_category = None
    current_level = None
    current_attribute = None

    # 标记是否在表头行
    for r in range(1, ws.max_row + 1):
        # 检查是否为表头行
        cell1 = ws.cell(r, 1).value
        if cell1 == "灵饰部位":
            continue  # 跳过表头行

        # 读取整行（Excel有12列）
        row = [ws.cell(r, c).value for c in range(1, 13)]

        # 更新当前值
        if row[0]:  # 灵饰部位
            current_position = row[0]
        if row[1]:  # 灵饰类别
            current_category = row[1]
        if row[2] is not None:  # 灵饰等级
            current_level = row[2]
        if row[3]:  # 属性
            current_attribute = row[3]

        # 解析左属性（第5-7列）
        left_desc = row[4]  # 第5列：部位/属性（如"戒指/伤害／3伤害"）
        left_price = parse_price(row[5])  # 第6列：价格
        left_boom_price = parse_price(row[6])  # 第7列：爆属性价格

        # 解析右属性（第8-10列）
        right_desc = row[7]  # 第8列：部位/属性
        right_price = parse_price(row[8])  # 第9列：价格
        right_boom_price = parse_price(row[9])  # 第10列：爆属性价格

        # 只有当左属性存在时才创建记录
        if left_desc:
            # 提取类型
            left_type = extract_type_from_desc(left_desc)
            right_type = extract_type_from_desc(right_desc) if right_desc else ''

            # 清理描述
            left_desc_clean = clean_desc(left_desc)
            right_desc_clean = clean_desc(right_desc) if right_desc else ''

            # 构建数据对象
            inferred_type = infer_type(current_position, current_category, current_attribute)
            final_type = left_type or right_type or inferred_type or ''

            item = {
                "position": current_position,
                "type": final_type,
                "category": current_category,
                "level": str(current_level) if current_level else '',
                "leftAttribute": left_desc_clean,
                "leftPrice": left_price,
                "leftBoomPrice": left_boom_price,
                "rightAttribute": right_desc_clean,
                "rightPrice": right_price,
                "rightBoomPrice": right_boom_price,
                "rawAttribute": current_attribute  # 原始属性字段
            }
            all_data.append(item)

    print(f"解析出 {len(all_data)} 条记录")

    # 统计数据
    positions = set([d["position"] for d in all_data])
    categories = set([d["category"] for d in all_data])
    levels = set([d["level"] for d in all_data])

    print(f"部位: {sorted(positions)}")
    print(f"类别: {sorted(categories)}")
    print(f"等级: {sorted(levels)}")

    # 生成JS文件
    js_content = f"""// 灵饰价格规则库
// 自动生成自: {os.path.basename(path)}
// 生成时间: {sys.version}
// 记录数: {len(all_data)}

const LINGSHI_DATA = {json.dumps(all_data, ensure_ascii=False, indent=2)};

// 导出供其他模块使用
if (typeof module !== 'undefined' && module.exports) {{
    module.exports = LINGSHI_DATA;
}}
"""

    with open(output_js_path, 'w', encoding='utf-8') as f:
        f.write(js_content)

    print(f"\n已生成JS数据文件: {output_js_path}")

    # 打印示例
    print("\n前5条记录示例:")
    for i, item in enumerate(all_data[:5]):
        print(f"  {i+1}: {item}")

if __name__ == "__main__":
    # 尝试多个可能的文件路径
    possible_paths = [
        r"E:\gujia\assets\sources\lingshi.xlsx",  # 根目录
        r"E:\Users\Administrator\Desktop\E:\gujia\assets\sources\lingshi.xlsx",  # 桌面
    ]

    excel_path = None
    for path in possible_paths:
        if os.path.exists(path):
            excel_path = path
            print(f"找到Excel文件: {excel_path}")
            break

    if not excel_path:
        print(f"错误: 找不到E:\gujia\assets\sources\lingshi.xlsx")
        print(f"请将文件放置在以下位置之一:")
        for path in possible_paths:
            print(f"  - {path}")
        sys.exit(1)

    output_path = r"E:\gujia\assets\data\lingshi_data.js"
    convert_excel_to_js(excel_path, output_path)