#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
调试宝宝价格表结构
"""

import openpyxl
import sys
import os

def debug_excel(file_path):
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    sheet_name = wb.sheetnames[0]
    ws = wb[sheet_name]

    print("行号,列1,列2,列3,列4,列5,列6,列7,列8,列9,列10,列11,列12,列13,列14")

    for r in range(1, min(ws.max_row, 200) + 1):
        row_vals = []
        for c in range(1, 15):  # 只看前14列
            val = ws.cell(r, c).value
            if val is None:
                row_vals.append('')
            else:
                row_vals.append(str(val).strip())

        # 检查是否有非空值
        if any(v != '' for v in row_vals):
            print(f"{r},{','.join(row_vals)}")

    wb.close()

if __name__ == "__main__":
    excel_file = r"E:\gujia\assets\sources\baobao.xlsx"
    debug_excel(excel_file)