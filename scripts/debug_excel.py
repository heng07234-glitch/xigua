#!/usr/bin/env python3
# -*- coding: utf-8 -*-
import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8')

def debug_excel_structure(path):
    print(f"调试Excel文件: {path}")

    wb = openpyxl.load_workbook(path, data_only=True)
    print(f"工作表数量: {len(wb.sheetnames)}")
    print(f"工作表名称: {wb.sheetnames}")

    for i, sheet_name in enumerate(wb.sheetnames):
        print(f"\n=== 工作表 {i}: {sheet_name} ===")
        ws = wb[sheet_name]
        print(f"行数: {ws.max_row}, 列数: {ws.max_column}")

        # 显示前10行
        print("前10行数据:")
        for r in range(1, min(11, ws.max_row + 1)):
            row_data = []
            for c in range(1, min(11, ws.max_column + 1)):  # 只检查前10列
                cell = ws.cell(r, c)
                if cell.value is not None:
                    row_data.append(f"{c}:{repr(cell.value)}")
            if row_data:
                print(f"  行{r}: {row_data}")

        # 检查第1行第1列的值
        first_cell = ws.cell(1, 1).value
        print(f"第1行第1列值: {repr(first_cell)}")

if __name__ == "__main__":
    excel_path = r"E:\gujia\assets\sources\lingshi.xlsx"
    debug_excel_structure(excel_path)