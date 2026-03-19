#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
从Excel估价表导入角色估价规则数据到MongoDB
将替换cbg.xigua集合中的现有数据
"""

import json
import sys
import os
import re
from datetime import datetime

# 尝试导入pymongo
try:
    from pymongo import MongoClient
except ImportError:
    print("错误: 未安装pymongo库")
    print("请运行: py -m pip install pymongo")
    sys.exit(1)

# 尝试导入openpyxl
try:
    import openpyxl
except ImportError:
    print("错误: 未安装openpyxl库")
    print("请运行: py -m pip install openpyxl")
    sys.exit(1)

def parse_price(price_str, is_base_price=True):
    """解析价格字符串，支持多种格式
    is_base_price: True表示基础价格（通常以"万"为单位），False表示附加价格（通常以"元"为单位）
    """
    if not price_str or str(price_str).strip() == '':
        return None

    price_str = str(price_str).strip()

    # 移除逗号分隔符
    price_str = price_str.replace(',', '')

    # 处理范围格式（如1000/1500）
    if '/' in price_str:
        parts = price_str.split('/')
        try:
            nums = [float(p.strip()) for p in parts if p.strip()]
            if nums:
                avg_price = sum(nums) / len(nums)  # 取平均值
                # 对于基础价格，如果平均值小于1000，可能需要乘以10000
                if is_base_price and avg_price < 1000 and not any('w' in p.lower() or '万' in p for p in parts):
                    avg_price = avg_price * 10000
                return avg_price
        except:
            return None

    # 检查是否包含单位
    original_str = price_str
    unit_multiplier = 1

    # 处理带单位的数值
    if price_str.lower().endswith('w') or price_str.endswith('万'):
        unit_multiplier = 10000
        price_str = price_str[:-1] if price_str.lower().endswith('w') else price_str[:-1]
    elif price_str.lower().endswith('k') or price_str.endswith('千'):
        unit_multiplier = 1000
        price_str = price_str[:-1] if price_str.lower().endswith('k') else price_str[:-1]

    try:
        price = float(price_str) * unit_multiplier

        # 特殊逻辑：对于基础价格，如果数值小于1000且没有明确单位，默认视为"万"单位
        if is_base_price and unit_multiplier == 1 and price < 1000:
            # 检查是否可能已经是实际价格（如"500"可能是500元）
            # 但根据Excel数据，基础价格如"18"表示18万=180,000元
            # 附加价格如"500"表示500元
            price = price * 10000

        # 对于附加价格，如果数值较大（>10000），可能也需要调整
        elif not is_base_price and unit_multiplier == 1 and price > 10000:
            # 可能是误将万单位解析为实际价格，但附加价格通常较小
            # 保持原值
            pass

        return price
    except:
        # 尝试处理特殊格式
        try:
            # 尝试直接转换
            return float(price_str)
        except:
            return None

def import_character_rules(excel_path, host="localhost", port=27017, database="cbg", collection="xigua"):
    """从Excel导入角色估价规则数据"""

    # Excel文件路径
    if not os.path.exists(excel_path):
        print(f"错误: Excel文件不存在: {excel_path}")
        return False

    print(f"正在读取Excel文件: {excel_path}")

    try:
        # 打开Excel文件
        wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)

        # 检查是否有"角色"工作表
        if '角色' not in wb.sheetnames:
            print(f"错误: Excel文件中没有'角色'工作表")
            print(f"可用的工作表: {wb.sheetnames}")
            wb.close()
            return False

        ws = wb['角色']

        # 解析数据
        rules_data = []
        current_level = None
        row_num = 0

        for row in ws.iter_rows(values_only=True):
            row_num += 1
            # 跳过完全空的行
            if all(cell is None or str(cell).strip() == '' for cell in row):
                continue

            # 转换行数据为列表
            row_cells = list(row)

            # 检查是否是等级标题行（第一列有数字）
            if row_cells[0] and str(row_cells[0]).strip().isdigit():
                current_level = str(row_cells[0]).strip()
                print(f"发现等级: {current_level}")
                continue

            # 检查是否是数据行（第二列有内容）
            if len(row_cells) > 1 and row_cells[1] and str(row_cells[1]).strip():
                type_str = str(row_cells[1]).strip()

                # 解析价格（第三列） - 基础价格
                price_str = row_cells[2] if len(row_cells) > 2 else None
                base_price = parse_price(price_str, is_base_price=True)

                # 解析限量物品（第五列）
                limit_item = None
                if len(row_cells) > 4 and row_cells[4]:
                    limit_item = str(row_cells[4]).strip()
                    if limit_item in ['', '无', 'None', 'null']:
                        limit_item = None

                # 解析附加价格（第六列） - 附加价格
                extra_price_str = row_cells[5] if len(row_cells) > 5 else None
                extra_price = parse_price(extra_price_str, is_base_price=False)

                # 计算总价格
                total_price = None
                if base_price is not None:
                    total_price = base_price
                    if extra_price is not None:
                        total_price += extra_price

                # 构建文档
                doc = {
                    'level': current_level,
                    'type': type_str,
                    'school': None,  # Excel中没有门派信息
                    'base_price': base_price,
                    'limit_item': limit_item,
                    'extra_price': extra_price,
                    'total_price': total_price,
                    'notes': '',  # 可以后续添加备注
                    'import_source': 'excel_rules',
                    'import_time': datetime.now(),
                    'row_number': row_num
                }

                rules_data.append(doc)
                print(f"  解析: {current_level}级 - {type_str}, 价格: {base_price}, 限量: {limit_item}, 附加: {extra_price}")

        wb.close()

        if not rules_data:
            print("警告: 没有解析到任何数据")
            return False

        print(f"解析完成，共 {len(rules_data)} 条规则")

        # 连接到MongoDB
        print(f"正在连接MongoDB: {host}:{port}")
        client = MongoClient(host, port)
        db = client[database]

        # 如果集合不存在则创建
        if collection not in db.list_collection_names():
            print(f"创建集合: {collection}")

        coll = db[collection]

        # 清空现有数据
        print(f"清空集合中的现有数据...")
        result = coll.delete_many({})
        print(f"已删除 {result.deleted_count} 条现有记录")

        # 插入新数据
        print(f"插入 {len(rules_data)} 条新记录...")
        insert_result = coll.insert_many(rules_data)

        print(f"成功插入 {len(insert_result.inserted_ids)} 条记录")

        # 创建索引
        coll.create_index([("level", 1)])
        coll.create_index([("type", 1)])
        coll.create_index([("base_price", 1)])
        print("索引创建完成")

        client.close()

        # 同时导出为JSON文件（与现有导出脚本兼容）
        export_json(rules_data)

        return True

    except Exception as e:
        print(f"导入数据时出错: {e}")
        import traceback
        traceback.print_exc()
        return False

def export_json(rules_data):
    """导出数据为JSON文件（与现有页面兼容）"""

    # 两个输出位置（与export_character_data.py一致）
    output_dir1 = os.path.join(os.path.dirname(__file__), "..", "assets", "data")
    output_file1 = os.path.join(output_dir1, "character_data.json")

    output_dir2 = os.path.join(os.path.dirname(__file__), "..", "docs", "app", "character")
    output_file2 = os.path.join(output_dir2, "character_data.json")

    # 确保输出目录存在
    os.makedirs(output_dir1, exist_ok=True)
    os.makedirs(output_dir2, exist_ok=True)

    # 准备导出数据 - 深拷贝并处理非序列化类型
    export_data_list = []
    for i, doc in enumerate(rules_data):
        export_doc = {}
        for key, value in doc.items():
            # 跳过_id字段（页面不需要）
            if key == '_id':
                continue

            # 处理datetime对象
            if isinstance(value, datetime):
                export_doc[key] = value.isoformat()
            else:
                # 尝试处理其他类型
                try:
                    # 检查是否可JSON序列化
                    json.dumps(value)
                    export_doc[key] = value
                except (TypeError, ValueError):
                    # 无法序列化，转换为字符串
                    export_doc[key] = str(value)

        export_data_list.append(export_doc)

    # 构建导出数据结构
    export_data = {
        "metadata": {
            "export_time": datetime.now().isoformat(),
            "source": "excel_character_rules",
            "document_count": len(export_data_list)
        },
        "data": export_data_list
    }

    # 写入JSON文件
    with open(output_file1, 'w', encoding='utf-8') as f:
        json.dump(export_data, f, ensure_ascii=False, indent=2)

    with open(output_file2, 'w', encoding='utf-8') as f:
        json.dump(export_data, f, ensure_ascii=False, indent=2)

    print(f"数据已导出到两个位置:")
    print(f"  1. {output_file1}")
    print(f"  2. {output_file2}")
    print(f"文档数量: {len(export_data_list)}")

def main():
    """主函数"""
    import argparse

    parser = argparse.ArgumentParser(description="从Excel导入角色估价规则数据到MongoDB")
    parser.add_argument("--excel", default=r"E:\Users\Administrator\Desktop\估价表.xlsx",
                       help="Excel文件路径（默认: E:\\Users\\Administrator\\Desktop\\估价表.xlsx）")
    parser.add_argument("--host", default="localhost", help="MongoDB主机地址（默认: localhost）")
    parser.add_argument("--port", type=int, default=27017, help="MongoDB端口（默认: 27017）")
    parser.add_argument("--database", default="cbg", help="数据库名称（默认: cbg）")
    parser.add_argument("--collection", default="xigua", help="集合名称（默认: xigua）")
    parser.add_argument("--dry-run", action="store_true", help="只解析不导入（测试模式）")

    args = parser.parse_args()

    print("=" * 60)
    print("角色估价规则数据导入工具")
    print("=" * 60)
    print(f"Excel文件: {args.excel}")
    print(f"MongoDB: {args.host}:{args.port}/{args.database}.{args.collection}")

    if args.dry_run:
        print("\n[测试模式] 只解析数据，不导入数据库")

    try:
        if args.dry_run:
            # 测试模式：只解析数据
            if not os.path.exists(args.excel):
                print(f"错误: Excel文件不存在: {args.excel}")
                return

            wb = openpyxl.load_workbook(args.excel, read_only=True, data_only=True)
            if '角色' not in wb.sheetnames:
                print(f"错误: Excel文件中没有'角色'工作表")
                print(f"可用的工作表: {wb.sheetnames}")
                wb.close()
                return

            ws = wb['角色']
            print(f"\n工作表'{ws.title}'信息:")
            print(f"  最大行数: {ws.max_row}")
            print(f"  最大列数: {ws.max_column}")

            # 显示前几行
            print("\n前10行数据:")
            for i, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True), 1):
                print(f"  行{i}: {row}")

            wb.close()
        else:
            # 实际导入
            success = import_character_rules(
                excel_path=args.excel,
                host=args.host,
                port=args.port,
                database=args.database,
                collection=args.collection
            )

            if success:
                print("\n[成功] 导入完成")
                print("\n下一步:")
                print("  1. 打开浏览器访问角色估价页面")
                print("  2. 刷新页面查看更新后的数据")
                print("  3. 使用筛选器查看不同等级的角色")
            else:
                print("\n[失败] 导入失败")

    except Exception as e:
        print(f"\n错误: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    main()